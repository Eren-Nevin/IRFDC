import csv 
from openpyxl import Workbook, load_workbook

def write_ads_to_csv(output_path, my_dict_list):
    with open(output_path, 'w') as file:
        writer = csv.DictWriter(file, my_dict_list[0].keys())
        writer.writeheader()
        writer.writerows(my_dict_list)

def write_ads_to_excel(filename, ad_dict_list, incremental=False):

    # If list is empty, return
    if not ad_dict_list:
        return
    if incremental:
        wb = load_workbook(filename)
    else:
        wb = Workbook()


    ws = wb.active

    if not incremental:
        column_names = list(ad_dict_list[0].keys())
        ws.append(column_names)

        for ad_dict in ad_dict_list:
            try:
                ws.append(list(ad_dict.values()))
            except:
                # There is a problem with data, so we print it. Usually it is a bad
                # formed data (e.g. a list instead of a number,...)
                print(list(ad_dict.values()))
    else:
        for rowy, ad_dict in enumerate(ad_dict_list, start=ws.max_row + 1):
            for colx, value in enumerate(ad_dict.values(), start=1):
                ws.cell(column=colx, row=rowy, value=value)
    wb.save(filename)

